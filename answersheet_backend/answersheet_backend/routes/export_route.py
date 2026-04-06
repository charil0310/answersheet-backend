import csv
import os
from datetime import datetime
from flask import Blueprint, request, send_file, jsonify
from flask_jwt_extended import jwt_required, get_jwt_identity
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from services.statistics_service import StatisticsService
from models.exam import Exam
from models.answer_sheet import AnswerSheet
from models.student import Student
from models.question import Question

export_bp = Blueprint("export", __name__)

OUTPUT_FOLDER = "output"
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

# 样式配置
BORDER_STYLE = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
HEADER_FONT = Font(bold=True, size=12)
ALIGN_CENTER = Alignment(horizontal='center', vertical='center')


@export_bp.route("/ranking/<int:exam_id>", methods=["GET"])
@jwt_required()
def export_ranking(exam_id):
    """
    导出学生成绩排名（Excel）
    ---
    tags:
      - Export
    parameters:
      - name: exam_id
        in: path
        required: true
        type: integer
    responses:
      200:
        description: 导出成功
      404:
        description: 考试不存在
    """
    teacher_id = int(get_jwt_identity())
    exam = Exam.query.filter_by(id=exam_id, teacher_id=teacher_id).first()
    if not exam:
        return jsonify({"error": "考试不存在或无权限"}), 404

    # 获取排名数据
    data = StatisticsService.ranking(exam_id)
    if not data:
        return jsonify({"error": "暂无成绩数据"}), 400

    # 创建Excel文件
    filename = f"排名_{exam.exam_name}_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    filepath = os.path.join(OUTPUT_FOLDER, filename)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "成绩排名"

    # 写入表头
    headers = ["排名", "学生姓名", "学号", "成绩"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_STYLE

    # 写入数据
    for row, item in enumerate(data, 2):
        ws.cell(row=row, column=1, value=item["rank"]).border = BORDER_STYLE
        ws.cell(row=row, column=2, value=item["student_name"]).border = BORDER_STYLE
        ws.cell(row=row, column=3, value=item["student_no"]).border = BORDER_STYLE
        ws.cell(row=row, column=4, value=item["score"]).border = BORDER_STYLE

    # 调整列宽
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 10

    wb.save(filepath)
    return send_file(filepath, as_attachment=True, download_name=filename)


@export_bp.route("/scores/<int:exam_id>", methods=["GET"])
@jwt_required()
def export_scores(exam_id):
    """
    导出学生成绩详情（Excel）
    ---
    tags:
      - Export
    parameters:
      - name: exam_id
        in: path
        required: true
        type: integer
    responses:
      200:
        description: 导出成功
      404:
        description: 考试不存在
    """
    teacher_id = int(get_jwt_identity())
    exam = Exam.query.filter_by(id=exam_id, teacher_id=teacher_id).first()
    if not exam:
        return jsonify({"error": "考试不存在或无权限"}), 404

    # 获取所有答题卡和学生信息
    sheets = AnswerSheet.query.filter_by(exam_id=exam_id).all()
    if not sheets:
        return jsonify({"error": "暂无答题卡数据"}), 400

    # 创建Excel文件
    filename = f"成绩详情_{exam.exam_name}_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    filepath = os.path.join(OUTPUT_FOLDER, filename)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "学生成绩"

    # 写入表头
    headers = ["学生姓名", "学号", "总分", "正确题数", "错误题数", "答题卡状态", "扫描时间"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_STYLE

    # 写入数据
    row = 2
    for sheet in sheets:
        student = Student.query.get(sheet.student_id) if sheet.student_id else None
        ws.cell(row=row, column=1, value=student.name if student else "未知").border = BORDER_STYLE
        ws.cell(row=row, column=2, value=student.student_no if student else "未知").border = BORDER_STYLE
        ws.cell(row=row, column=3, value=sheet.total_score or 0).border = BORDER_STYLE
        ws.cell(row=row, column=4, value=sheet.correct_count or 0).border = BORDER_STYLE
        ws.cell(row=row, column=5, value=sheet.wrong_count or 0).border = BORDER_STYLE
        ws.cell(row=row, column=6, value=sheet.status).border = BORDER_STYLE
        scan_time = sheet.scan_time.strftime("%Y-%m-%d %H:%M:%S") if sheet.scan_time else ""
        ws.cell(row=row, column=7, value=scan_time).border = BORDER_STYLE
        row += 1

    # 调整列宽
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 20

    wb.save(filepath)
    return send_file(filepath, as_attachment=True, download_name=filename)


@export_bp.route("/statistics/<int:exam_id>", methods=["GET"])
@jwt_required()
def export_statistics(exam_id):
    """
    导出考试统计分析（Excel）
    ---
    tags:
      - Export
    parameters:
      - name: exam_id
        in: path
        required: true
        type: integer
    responses:
      200:
        description: 导出成功
      404:
        description: 考试不存在
    """
    teacher_id = int(get_jwt_identity())
    exam = Exam.query.filter_by(id=exam_id, teacher_id=teacher_id).first()
    if not exam:
        return jsonify({"error": "考试不存在或无权限"}), 404

    # 获取统计数据
    basic_stats = StatisticsService.basic_stats(exam_id)
    distribution = StatisticsService.score_distribution(exam_id)
    accuracy = StatisticsService.question_accuracy(exam_id)

    # 创建Excel文件
    filename = f"统计分析_{exam.exam_name}_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    filepath = os.path.join(OUTPUT_FOLDER, filename)

    wb = openpyxl.Workbook()

    # 1. 基础统计sheet
    ws1 = wb.active
    ws1.title = "基础统计"
    ws1.cell(row=1, column=1, value="考试名称").font = HEADER_FONT
    ws1.cell(row=1, column=2, value=exam.exam_name)
    ws1.cell(row=2, column=1, value="科目").font = HEADER_FONT
    ws1.cell(row=2, column=2, value=exam.course_name or "未设置")
    ws1.cell(row=3, column=1, value="平均分").font = HEADER_FONT
    ws1.cell(row=3, column=2, value=basic_stats["average"])
    ws1.cell(row=4, column=1, value="最高分").font = HEADER_FONT
    ws1.cell(row=4, column=2, value=basic_stats["max"])
    ws1.cell(row=5, column=1, value="最低分").font = HEADER_FONT
    ws1.cell(row=5, column=2, value=basic_stats["min"])

    # 2. 分数分布sheet
    ws2 = wb.create_sheet("分数分布")
    headers = ["分数段", "人数"]
    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_STYLE

    row = 2
    for range_name, count in distribution.items():
        ws2.cell(row=row, column=1, value=range_name).border = BORDER_STYLE
        ws2.cell(row=row, column=2, value=count).border = BORDER_STYLE
        row += 1

    # 3. 每题正确率sheet
    ws3 = wb.create_sheet("每题正确率")
    headers = ["题号", "正确率"]
    for col, header in enumerate(headers, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_STYLE

    row = 2
    for item in accuracy:
        ws3.cell(row=row, column=1, value=item["question_no"]).border = BORDER_STYLE
        ws3.cell(row=row, column=2, value=f"{item['accuracy']*100:.2f}%").border = BORDER_STYLE
        row += 1

    wb.save(filepath)
    return send_file(filepath, as_attachment=True, download_name=filename)


@export_bp.route("/comparison", methods=["POST"])
@jwt_required()
def export_comparison():
    """
    导出多次考试对比（Excel）
    ---
    tags:
      - Export
    parameters:
      - name: body
        in: body
        required: true
        schema:
          type: object
          properties:
            exam_ids:
              type: array
              items:
                type: integer
              example: [1,2,3]
    responses:
      200:
        description: 导出成功
      400:
        description: 参数错误
    """
    teacher_id = int(get_jwt_identity())
    data = request.json
    exam_ids = data.get("exam_ids", [])
    if not exam_ids:
        return jsonify({"error": "请选择要对比的考试"}), 400

    # 验证考试权限
    exams = []
    for exam_id in exam_ids:
        exam = Exam.query.filter_by(id=exam_id, teacher_id=teacher_id).first()
        if exam:
            exams.append(exam)
    if not exams:
        return jsonify({"error": "无可用的考试数据"}), 400

    # 创建Excel文件
    filename = f"考试对比_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    filepath = os.path.join(OUTPUT_FOLDER, filename)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "多次考试对比"

    # 写入表头
    headers = ["考试名称", "科目", "平均分", "最高分", "最低分", "参考人数"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_STYLE

    # 写入每个考试的统计数据
    row = 2
    for exam in exams:
        basic = StatisticsService.basic_stats(exam.id)
        sheet_count = AnswerSheet.query.filter_by(exam_id=exam.id).count()
        
        ws.cell(row=row, column=1, value=exam.exam_name).border = BORDER_STYLE
        ws.cell(row=row, column=2, value=exam.course_name or "未设置").border = BORDER_STYLE
        ws.cell(row=row, column=3, value=basic["average"]).border = BORDER_STYLE
        ws.cell(row=row, column=4, value=basic["max"]).border = BORDER_STYLE
        ws.cell(row=row, column=5, value=basic["min"]).border = BORDER_STYLE
        ws.cell(row=row, column=6, value=sheet_count).border = BORDER_STYLE
        row += 1

    wb.save(filepath)
    return send_file(filepath, as_attachment=True, download_name=filename)